import openpyxl, csv, os

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)
        for sheetName in wb.get_sheet_names():
            # Loop trough every sheet in the workbook.
            sheet = wb.get_sheet_by_name(sheetName)

            # Create the CSV filename from the Excel filename and sheet title.
            csvFile = open(str(excelFile.strip('xlsx')) + '_' + str(sheetName) + '.csv', 'w', newline ='')
            # Create the csv.writer object for this CSV file.
            csvWriter = csv.writer(csvFile)

            # Loop trough every row in the sheet.
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []    # append each cell to this list
                # Loop trough each cell in the row.
                for colNum in range(1, sheet.max_column + 1):
                    # Append each cell's data to rowData.
                    cell = sheet.cell(row=1, column=colNum)
                    rowData.append(cell.value)
                # Write the rowData list to the CSV file.
                print('Converting ' + str(excelFile) + ' to .csv...')
                for i in rowData:
                    csvWriter.writerow([i])
            csvFile.close()
print('Done.')            